import logging
from os import system
import sqlite3
import time
from sqlite3 import DatabaseError
from uu import Error
import bcrypt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
from django.db import connections,transaction
from system import models
from system.models import PersonalBankAccount,CashTransaction,CashInHand,FDAccountModel, RDAccountModel
from django.db.models import F
from django.http import HttpResponse
import math

logging.basicConfig(
    level=logging.DEBUG,  # Set the minimum logging level
    format="%(asctime)s - %(levelname)s - %(message)s"  # Define the format of log messages
)

from datetime import datetime


def get_current_date():
    # Get the current date
    current_date = datetime.now().date()

    # Format the current date as 'yy-mm-dd'
    formatted_date = current_date.strftime('%Y-%m-%d')

    return formatted_date


def add_days_to_date(date_string, days_to_add):
    try:
        # Parse the date string to get year, month, and day
        year, month, day = map(int, date_string.split('-'))

        # Create a datetime object for the input date
        input_date = datetime(year, month, day)

        # Add the specified number of days to the input date
        result_date = input_date + timedelta(days=days_to_add)

        # Format the result date as a string in the format 'yy-mm-dd'
        return result_date.strftime('%Y-%m-%d')
    except ValueError:
        # Handle invalid date format
        logging.error("Invalid date format. Expected format: YY-MM-DD")
        return None


def ac_availability(ann):
    try:
        database_name='other_database'
        with connections[database_name].cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM personal_bank_account WHERE account_number = %s", [ann])
            row = cursor.fetchone()
            if row[0] > 0:
                return True
            else:
                return False
            
    except Exception as e:
        # Handle any exceptions (e.g., database connection errors)
        logging.warning("Error while checking account availability:", e)
        return False


def download_sta():
    an7 = int(input("Enter account number: "))
    if ac_availability(an7):
        conn = sqlite3.connect('bank_manage.db')
        query = "SELECT * FROM cash_transaction WHERE ac_no=?"
        df = pd.read_sql_query(query, conn, params=(an7,))
        # Close the database connection
        conn.close()
        wb = Workbook()
        ws = wb.active
        # Write DataFrame to the worksheet starting from cell A1
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        # Apply formatting to the header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Apply center alignment to all cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # Save the workbook to a new Excel file
        wb.save('account_statement.xlsx')
        # Log a message indicating that the Excel file is downloaded
        logging.info("account statement is downloaded")
    else:
        print("Account number", an7, "is not found please try again!...")



def save_transaction(account=None, transaction_t=None, dt7=None, a1=None, cih=0, vn=None, fan=0, tan=0, user=None):
    try:
        print("Initializing save transaction")
        if cih is None:
            raise ValueError("cash_in_hand_previous (cih) cannot be None")
        
        with transaction.atomic(using='other_database'):
            transaction = CashTransaction(
                ac_no=account,
                transaction_type=transaction_t,
                date=dt7,
                amt=a1,
                cash_in_hand_previous=cih,
                voucher_no=vn,
                frm_ac_no=fan,
                to_ac_no=tan,
                username=user
            )
        # Save the transaction object to the database specified by the 'other_database' alias
        transaction.save(using='other_database')
        print("Completed save transaction")
    except ValueError as ve:
        logging.error("ValueError in save_transaction: %s", ve)
    except KeyError as ke:
        logging.error("KeyError in save_transaction: %s", ke)
    except AttributeError as ae:
        logging.error("AttributeError in save_transaction: %s", ae)
    except TypeError as te:
        logging.error("TypeError in save_transaction: %s", te)
    except Exception as e:
        logging.error("Error occurred while saving transaction: %s", e)

def revoke_transaction():
    global res1

    def save_detail(an0, dt, amt1, tt, fan=None, tan=None, vn9=None):
        try:
            nc = sqlite3.connect('bank_manage.db')
            try:
                q4 = ("INSERT INTO revoke_history(ac_no,date,amt,tr_ty,from_ac_no,to_ac_no,voucher_no) VALUES (?,?,?,"
                      "?,?,?,?)")
                nc.execute(q4, (an0, dt, amt1, tt, fan, tan, vn9))
                nc.commit()
            except sqlite3.Error as e:
                logging.warning("Error while saving transaction detail: %s", e)
        except sqlite3.Error as e:
            logging.warning("Error while connecting to database: %s", e)
        finally:
            if 'nc' in locals():
                nc.close()

    print("TO revoke a transaction please authenticate with your account")
    un1 = input("enter your username: ")
    pw = input("enter your password: ")
    res = authenticate_user(un1, pw)

    if res:
        ip3 = int(input("Enter voucher no to Revoke: "))
        try:
            connection = sqlite3.connect("bank_manage.db")
            try:
                curr = connection.cursor()
                cq = "SELECT * FROM cash_transaction WHERE voucher_no=?"
                curr.execute(cq, (ip3,))
                res0 = curr.fetchone()

                try:
                    cn = sqlite3.connect("bank_manage.db")
                    try:
                        cur00 = cn.cursor()
                        q9 = "SELECT voucher_no FROM revoke_history WHERE ac_no=?"
                        cur00.execute(q9, (res0[0],))
                        res1 = cur00.fetchone()
                    except sqlite3.Error as e:
                        logging.warning("Error while getting info: %s", e)
                except sqlite3.OperationalError as e:
                    logging.critical("Error while connecting to database: %s", e)
                if res1[0] == ip3:
                    logging.info("The Transaction is already Revoked")
                else:
                    if res0[1] == "CASH DEPOSIT":
                        q = "UPDATE personal_bank_account SET balance=balance - ? WHERE account_number=?"
                        curr.execute(q, (res0[3], res0[0],))
                        connection.commit()
                        date = get_current_date()
                        save_detail(res0[0], date, res0[3], res0[1], vn9=ip3)
                        logging.info("Successfully Transaction is revoked")
                    elif res0[1] == "CASH WITHDRAW":
                        q = "UPDATE personal_bank_account SET balance=balance + ? WHERE account_number=?"
                        curr.execute(q, (res0[3], res0[0],))
                        dt = get_current_date()
                        save_detail(res0[0], dt, res0[3], res0[1], vn9=ip3)
                        connection.commit()
                        logging.info("Successfully Transaction is revoked")
                    elif res0[1] == "TRF":
                        logging.info("Successfully Transaction is revoked")
                        pass
                    else:
                        print("Unexpected error")
            except sqlite3.OperationalError as e:
                logging.warning("Error while getting transaction details: %s", e)
        except sqlite3.Error as e:
            logging.critical("Error while connecting to database: %s", e)
        finally:
            if 'connection' in locals():
                connection.close()


def get_voucher_no():
    database_name = 'other_database'  # Set the database name here
    try:
        with connections[database_name].cursor() as cursor:
            try:
                cursor.execute("SELECT voucher_no FROM other")
                vn1 = cursor.fetchone()
                return vn1[0] if vn1 else None
            except Exception as e:
                logging.error("Error fetching voucher number: %s", e)
                return None
    except Exception as e:
        logging.warning("Error connecting to the database: %s", e)
        return None

def create_ac(nm,at,ad,mn,an,pn,user):
    # it creates the account no in the database by using simple query INSERT
        with connections['other_database'].cursor() as cursor:
            try:
                dt = get_current_date()
                cursor.execute("""
                    INSERT INTO personal_bank_account
                    (account_holder_name, balance, account_type, opening_date, address, mobile_number, aadhar_card_number, pan_card_number,username)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?,?)
                """, (nm, 0, at, dt, ad, mn, an, pn,user))
                transaction.commit()
                cursor.execute("SELECT account_number FROM personal_bank_account ORDER BY account_number DESC LIMIT 1")
                account_number = cursor.fetchone()[0]
                print("...........Customer Account Created successfully........")
                return account_number
            except Exception as e:
                print(e)
                transaction.rollback(using='other_database')
            except sqlite3.Error as e:
                print(e)


class Customer:
    # This class is separately for Editing the customer personal information only
    # it gives parameter by using its constructor is only a account no.
    def __init__(self, an0,user):
        self.an = an0
        self.user=user

    def change_pan(self):
        pn = input("Enter pan card Number to Update")
        cono = sqlite3.connect("bank_manage.db")
        q = "UPDATE personal_bank_account SET pan_card_number=? WHERE account_number=?"
        cono.execute(q, (pn, self.an))
        cono.commit()
        cono.close()
        logging.info("Pan Card Number Updated Successfully")

    def change_card_number(self):
        pn = input("Enter aadhar card Number to Update")
        con = sqlite3.connect("bank_manage.db")
        q = "UPDATE personal_bank_account SET aadhar_card_number=? WHERE account_number=?"
        con.execute(q, (pn, self.an))
        con.commit()
        con.close()

        logging.info("Pan Card Number Updated Successfully")

    def change_mobile_no(self):
        mn = input("Enter Mobile card Number to Update")
        con = sqlite3.connect("bank_manage.db")
        q = "UPDATE personal_bank_account SET mobile_number=? WHERE account_number=?"
        con.execute(q, (mn, self.an))
        con.commit()
        logging.info("Mobile Number Updated Successfully")

    def get_balance(self):
        try:
            # Query the PersonalBankAccount table using the account_number and username
            account = PersonalBankAccount.objects.using('other_database').get(account_number=self.an, username=self.user)
            return account.balance  # Return the balance from the retrieved account
        except PersonalBankAccount.DoesNotExist:
            logging.warning("No balance found for account number %s and username %s", self.an, self.username)
            return None
        except Exception as e:
            logging.error("Error fetching balance: %s", e)
            return None



class Wallet:
    # This class is for only purpose of handling cash in hand transaction
    # this class get initial amt amount that we give by fetching from our database at the time of calling
    # Firstly its return the cash in hand or Another function set the cash in hand
    def __init__(self, set_amt,user,cih):
        self.set_amt = set_amt
        self.user=user
        self.cih=cih

    def cash_in_hand_return(self):
        return self.set_amt

    def cash_in_hand_handle (self):

        try:
            print("in cih")
            # Update the cash_in_hand field for the specified username
            query = """
                UPDATE cash_in_hand
                SET cash_in_hand = %s
                WHERE username = %s
            """
            # Execute the raw SQL query with the provided parameters
            print("connecting to database..")
            with connections['other_database'].cursor() as cursor:
                print("excuting query")
                cursor.execute(query, [self.set_amt, self.user])
                print("query excutes")
                transaction.commit()
                print("transaction complete")
            return True  # Return True if the update is successful
        except CashInHand.DoesNotExist:
            print("cash in hand does not exist")
            # Handle the case where the record does not exist for the provided username
            return False
        except Exception as e:
            # Handle other exceptions
            print("Error updating cash_in_hand:", e)
            return False      


class Account:
    # This is major class of this program it's having some main member functions are included in it this class
    # inherits the class Wallet in it Class Wallet have an one argument called initial Class Account have one
    # argument called an Account number Because of this argument or Inheritance we have to call a base class
    # constructor manually that why pythons allowed to call a manual constructor
    def __init__(self, an2, amt6,user):
        self.amt2 = amt6
        self.an2 = an2
        self.user=user

    def Deposit(self):
        database_name = 'other_database'  
        try:
            with transaction.atomic(using=database_name):
                with connections[database_name].cursor() as cursor:
                    try:
                        q1 = "UPDATE personal_bank_account SET balance=balance + %s WHERE account_number = %s"
                        Q2 = "UPDATE other SET voucher_no=voucher_no+1"
                        cursor.execute(q1, (self.amt2, self.an2))
                        cursor.execute(Q2)
                        try:
                            # getting cash in hand details by orm
                            print("getting cash in hand ")
                            cash_in_hand_instance = CashInHand.objects.using('other_database').get(username=self.user)
                            cash_in_hand = int(cash_in_hand_instance.cash_in_hand)
                            print(cash_in_hand)
                            print("Seeting amt")
                            set_amt=int(cash_in_hand)+int(self.amt2)
                            print("amt set and is:",set_amt)
                            try:
                                print("creating obj of the wallet")
                                w1=Wallet(set_amt=set_amt,user=self.user,cih=cash_in_hand)
                                print("calling cash in handle")
                                v=w1.cash_in_hand_handle()
                                vn = get_voucher_no()

                                # this is for only saving the transactions purpose only
                                if vn is not None:
                                    logging.info("Amount Credited To AC NO.%s voucher no is.%s", self.an2, vn + 1)
                                    print("creating obj of customer")
                                    ob = Customer(self.an2,self.user)
                                    print("getting balance")
                                    ab = ob.get_balance()
                                    print("getting date")
                                    dt = get_current_date()
                                    tt = "CASH WITHDRAWL"
                                    print("calling save transaction")
                                    save_transaction(account=int(self.an2), transaction_t=str(tt), dt7=str(dt), a1=int(self.amt2), cih=int(cash_in_hand), vn=int(vn+1), user=str(self.user) )
                                    print("called completet to save transaction")
                                    return True,int(vn+1)
                                else:
                                    logging.warning("Error while fetching voucher no.")
                            except Exception as e:
                                transaction.rollback(using=database_name)
                                print("error while calling cash in handle",e)
                        except Exception as e:
                                    transaction.rollback(using=database_name)
                                    print(e)
                                    return HttpResponse("error while getting cash in details",e)
                    except Exception as e:
                        transaction.rollback(using=database_name)
                        logging.error("Transaction failed: %s", e)
        except Exception as e:
            # Rollback the transaction in case of error
            transaction.rollback(using=database_name)
            logging.error("Transaction failed: %s", e)
            return HttpResponse("Transaction failed: {}".format(e))
                
                   

    def withdraw(self):
        print("in deposit")
        database_name = 'other_database'  
        try:
            with transaction.atomic(using=database_name):
                with connections[database_name].cursor() as cursor:
                    try:
                        q1 = "UPDATE personal_bank_account SET balance=balance - %s WHERE account_number = %s"
                        Q2 = "UPDATE other SET voucher_no=voucher_no+1"
                        cursor.execute(q1, (self.amt2, self.an2))
                        cursor.execute(Q2)
                        try:
                            # getting cash in hand details by orm
                            print("getting cash in hand ")
                            cash_in_hand_instance = CashInHand.objects.using('other_database').get(username=self.user)
                            cash_in_hand = int(cash_in_hand_instance.cash_in_hand)
                            print(cash_in_hand)
                            print("Seeting amt")
                            set_amt=int(cash_in_hand)-int(self.amt2)
                            print("amt set and is:",set_amt)
                            try:
                                print("creating obj of the wallet")
                                w1=Wallet(set_amt=set_amt,user=self.user,cih=cash_in_hand)
                                print("calling cash in handle")
                                v=w1.cash_in_hand_handle()
                                vn = get_voucher_no()

                                # this is for only saving the transactions purpose only
                                if vn is not None:
                                    logging.info("Amount Debited From AC NO.%s voucher no is.%s", self.an2, vn + 1)
                                    print("creating obj of customer")
                                    ob = Customer(self.an2,self.user)
                                    print("getting balance")
                                    ab = ob.get_balance()
                                    print("getting date")
                                    dt = get_current_date()
                                    tt = "CASH WITHDRAWL"
                                    print("calling save transaction")
                                    save_transaction(account=int(self.an2), transaction_t=str(tt), dt7=str(dt), a1=int(self.amt2), cih=int(cash_in_hand), vn=int(vn+1), user=str(self.user) )
                                    print("called completet to save transaction")
                                    return True,int(vn+1)
                                else:
                                    logging.warning("Error while fetching voucher no.")
                            except Exception as e:
                                transaction.rollback(using=database_name)
                                print("error while calling cash in handle",e)
                        except Exception as e:
                                    transaction.rollback(using=database_name)
                                    print(e)
                                    return HttpResponse("error while getting cash in details",e)
                    except Exception as e:
                        transaction.rollback(using=database_name)
                        logging.error("Transaction failed: %s", e)
        except Exception as e:
            # Rollback the transaction in case of error
            transaction.rollback(using=database_name)
            logging.error("Transaction failed: %s", e)
            return HttpResponse("Transaction failed: {}".format(e))    

    def transfer(self, a1,a2):
        self.amt2=int(self.amt2)
        global rr, cc, cc2
        print(a1,a2,type(a1),type(a2))
        database_name='other_database'
        with transaction.atomic(using=database_name):
            with connections[database_name].cursor() as cursor:
                print("connection suceed to database")
                try:
                    qq = "SELECT balance FROM personal_bank_account WHERE account_number = {}".format(a1)
                    cursor.execute(qq)
                    print("Account fetching query has been successfully executed")
                    rr = cursor.fetchone()
                    print(rr)
                except sqlite3.Error as e:
                    print("Error while getting details...", e)
                if int(rr[0]) > int(self.amt2):
                    qq1 = "UPDATE personal_bank_account SET balance=balance-{} WHERE account_number={}".format(int(self.amt2), a1)
                    qq2 = "UPDATE personal_bank_account SET balance=balance+{} WHERE account_number={}".format(int(self.amt2), a2)
                    try:
                        cursor.execute(qq1)
                        print("first updatin query is done")
                        try:
                            cursor.execute(qq2)
                            print('second query has been done')
                            tt = "TRF"
                            d = get_current_date()
                            vn = get_voucher_no()
                            print("calling to save transaction")
                            save_transaction(transaction_t=tt, dt7=d, vn=vn, fan=a1, tan=a2,user=self.user)  # Calling the function
                            print("%s RS. successfully Transferred to account number: %s", self.amt2, a2)
                            return True
                        except sqlite3.Error as e:
                            transaction.rollback(using=database_name)
                            logging.warning("Transfer failed due to", e)
                            print("save tra",e)
                            return False
                    except:
                        transaction.rollback(using=database_name)
                        logging.warning("Transaction failed....", exc_info=True)
                        print("main query",e)
                else:
                    logging.info("The Account Balance Is Too Low To Debit The Amt")
                    print("The Account Balance Is Too Low To Debit The Amt")
                    return "balance low"
                

    def apply_interest(self, intrest, period):
        pass

    def close_ac(self):
        an7 = int(input("Enter account number to close"))
        try:
            cc3 = sqlite3.connect('bank_manage.db')
            try:
                cc0 = cc3.cursor()
                q = "DELETE FROM personal_bank_account WHERE ac_no=?"
                qq2 = "SELECT * FROM personal_bank_account WHERE ac_no=?"
                result0 = cc0.execute(qq2, (an7,))
                cc3.commit()
                try:
                    cc0.execute(q, (an7,))
                    cc3.commit()
                except:
                    logging.info("Error while getting to save details")
            except sqlite3.Error as e:
                logging.info("Error while trying to get details")
        except:
            logging.info("Error while trying to connect to database")
        finally:
            cc3.close()


class Transaction:
    pass


class Common_functions:
    def __init__(self, AMT, cr_ac, debit_ac):
        self.AMT = AMT
        self.cr_ac = cr_ac
        self.debit_ac = debit_ac

    def transfer_funds(self):
        try:
            # Connect to the SQLite database
            connection = sqlite3.connect("bank_manage.db")
            cursor = connection.cursor()
            # Begin transaction
            cursor.execute("BEGIN TRANSACTION")
            # Deduct amount from debit account
            cursor.execute("UPDATE personal_bank_account SET balance = balance - ? WHERE account_number = ?",
                           (self.AMT, self.debit_ac))
            # Add amount to credit account
            cursor.execute("UPDATE fd_accounts SET account_balance = ? WHERE fd_ac_no = ?",
                           (self.AMT, self.cr_ac))
            connection.commit()
            # save to the sql table
            save_transaction(transaction_t="TRF", dt7=get_current_date(), vn=get_voucher_no(), fan=self.debit_ac,
                             tan=self.cr_ac)
        except sqlite3.DatabaseError as e:
            logging.error(f"Database error: {e}")
            # Rollback transaction if an error occurs
            connection.rollback()
        except Exception as e:
            logging.error(f"Error: {e}")
            # Rollback transaction if an error occurs
            connection.rollback()
        finally:
            # Close the database connection
            if connection:
                connection.close()

    def self_bank_transfer(self):
        pass


class Fdaccount:
    def __init__(self, an88=None):
        self.an88 = an88

    def get_account_details(self, clause):
        try:
            c = sqlite3.connect("bank_manage.db")
            try:
                # Getting all details form fd_ac_table information for specific account number
                cu = c.cursor()
                query8 = "SELECT * FROM fd_accounts WHERE fd_ac_no=?"
                cu.execute(query8, (clause,))
                rows = cu.fetchone()
                return rows
            except sqlite3.Error as e:
                logging.info("Error while getting information from database", e)
        except sqlite3.OperationalError as e:
            logging.warning(e)
        finally:
            if c:
                c.close()

    def check_account(self, ann):
        temp = False
        try:
            c2 = sqlite3.connect("bank_manage.db")
            try:
                cc5 = c2.cursor()
                q6 = "SELECT fd_ac_no FROM fd_accounts"
                cc5.execute(q6)
                r9 = cc5.fetchall()
                print(r9)
                for row in r9:
                    if row[0] == ann:
                        temp = True
                        break
                return temp
            except sqlite3.Error as e:
                logging.info("Error while getting details", e)
        except:
            logging.warning("Error while connecting to database")
        finally:
            if c2 in locals():
                c2.close()

    def get_account_number(self):
        try:
            cc30 = sqlite3.connect('bank_manage.db')
            try:
                cc30c = cc30.cursor()
                q01 = "SELECT fd_ac_no FROM fd_accounts"
                cc30c.execute(q01)
                r9 = cc30c.fetchall()
                return r9[-1]
            except Exception as e:
                logging.info("Error getting to get account number", e)
        except sqlite3.Error as e:
            logging.info("Error while connecting to database")

    def calculate_mat_amt(self, fd_amt, ir1, fd_days1):
        interest1 = (fd_amt * ir1 * fd_days1) / (365 * 100)
        return interest1

    def days_between_dates(self, date1_str, date2_str):
        # Convert the date strings to datetime objects
        date1 = datetime.strptime(date1_str, '%Y-%m-%d')
        date2 = datetime.strptime(date2_str, '%Y-%m-%d')
        delta = date2 - date1
        
        # Return the absolute value of the number of days
        return abs(delta.days)

    def create_account(self,user,personal_ac_no,fd_am,fd_days,ir):
        database_name='other_database'
        with transaction.atomic(using=database_name):
            with connections[database_name].cursor() as cursor:
                od = str(get_current_date())
                try:
                    customer=PersonalBankAccount.objects.using('other_database').get(account_number=personal_ac_no)
                    fd_intrest= (int(fd_am) * int(ir) * int(fd_days)) / (365 * 100)
                    fd_mat_amt=float(fd_intrest)+float(fd_am)
                    new_account = FDAccountModel.objects.using('other_database').create(
                    customer_name=customer.account_holder_name,
                    opening_date=od,
                    int_rate=ir,
                    fd_days=fd_days,
                    fd_opening_amt=fd_am,
                    personal_ac_no=personal_ac_no,
                    username=user,
                    account_balance=0,  # Assuming initial balance
                    renew=0 ,
                    pre_mature_withdraw=False,
                    mat_amt= fd_mat_amt,
                    fd_mat_dt=add_days_to_date(od, int(fd_days)) ,
                    )
                    new_account.save

                    print("Account Number created sucessfully")
                    return True  
                except Exception as r:
                    print("the error from the bank_management",r)
                    return False
                

                       
    def delete_account(self):
        pass

    def check_fd_balance(self, fd_ac_n):
            # check FD balance and FD opening balance and returns tuple
            try:
                o = sqlite3.connect("bank_manage.db")
                try:
                    c1 = o.cursor()
                    qq2 = "SELECT account_balance,fd_opening_amt,pre_mature_withdraw FROM fd_accounts WHERE fd_ac_no=?"
                    c1.execute(qq2, (fd_ac_n,))
                    r22 = c1.fetchone()
                    return r22
                except sqlite3.Error as r:
                    logging.info("Error while getting details", r)
            except:
                logging.error("Error while connecting to database")

    def add_funds(self, fd_obj, user, ac_no):

        database_name='other_database'         
        try:
            with transaction.atomic(using=database_name):
                with connections[database_name].cursor() as cursor:
                    try:
                        cih=CashInHand.objects.using(database_name).get(username=user)
                        set_amt=int(cih.cash_in_hand)+int(fd_obj.fd_opening_amt)
                        print(set_amt)
                        obj=Wallet(set_amt,user,int(cih.cash_in_hand))
                        obj.cash_in_hand_handle()
                        try:
                            record=FDAccountModel.objects.using(database_name).get(username=user,fd_ac_no=ac_no)
                            record.account_balance=fd_obj.fd_opening_amt
                            record.save()
                            print("fund added successfuly")
                            return True
                        except Exception as e:
                            print("error occured in when adding to fd balance", e)
                    except Exception as e:
                        print("error on whole transaction",e)
                        transaction.rollback(using=database_name)
        except Exception as e:
            transaction.rollback(using=database_name)
            print("error while connecting to database",e)


    def withdraw(self):
            fd_Ac_n = int(input("Enter the FD Account Number: "))
            if self.check_account(fd_Ac_n):
                fd_bal = self.check_fd_balance(fd_Ac_n)
                if fd_bal != 0:
                    try:
                        ccu = sqlite3.connect("bank_manage.db")
                        ccu1 = ccu.cursor()
                        qq8 = "SELECT * FROM fd_accounts WHERE fd_ac_no=?"
                        ccu1.execute(qq8, (fd_Ac_n,))
                        fd_res1 = ccu1.fetchone()
                        dt = str(get_current_date())
                        if dt == fd_res1[10]:
                            print("Maturity Date is Over. You can proceed to withdraw FD.")
                            if fd_res1[2] == fd_res1[8]:
                                try:
                                    q11 = "UPDATE fd_accounts SET account_balance=account_balance - ? WHERE fd_ac_no=?"
                                    ccu1.execute(q11, (fd_res1[2], fd_Ac_n))
                                    ccu.commit()
                                    withdraw = 1
                                    try:
                                        q12 = ("UPDATE personal_bank_account SET balance=balance + ?,pre_mature_withdraw=? "
                                            "WHERE account_number=?")
                                        ccu1.execute(q12, (fd_res1[7], withdraw, fd_res1[9]))
                                        ccu.commit()
                                        logging.info("Successfully transferred amount to Saving account")
                                    except Exception as e:
                                        ccu.rollback()
                                        logging.error("Transfer failed:", e)
                                except Exception as e:
                                    ccu.rollback()
                                    logging.error("Transfer failed to personal Account:", e)

                        else:
                            try:
                                c = int(
                                    input("Today is not maturity date. Please confirm to premature withdraw. Press 1: "))
                                if c == 1:
                                    try:
                                        ccu0 = sqlite3.connect("bank_manage.db")
                                        ccu1 = ccu0.cursor()

                                        ccu1.execute("SELECT * FROM fd_accounts WHERE fd_ac_no=?", (fd_Ac_n,))
                                        fd_account = ccu1.fetchone()

                                        current_date = str(get_current_date())
                                        days_diff = int(self.days_between_dates(current_date, fd_account[3]))
                                        if fd_account[2] == fd_account[8]:
                                            try:
                                                saving_account_update_query = ("UPDATE personal_bank_account SET "
                                                                            "balance=balance + ? WHERE account_number=?")
                                                ccu1.execute(saving_account_update_query, (
                                                    self.calculate_mat_amt(fd_account[2], fd_account[4], days_diff),
                                                    fd_account[9]))
                                                ccu0.commit()
                                                withdraw = 1

                                                fd_account_update_query = ("UPDATE fd_accounts SET "
                                                                        "account_balance=account_balance - ?,"
                                                                        "pre_mature_withdraw=? WHERE fd_ac_no=?")
                                                ccu1.execute(fd_account_update_query, (fd_account[2], withdraw, fd_Ac_n))
                                                ccu0.commit()
                                                logging.info("FD withdrawal successful. Amount transferred to Saving "
                                                            "Account.")
                                            except sqlite3.DatabaseError as e:
                                                ccu0.rollback()
                                                logging.error("Error occurred while updating database:", e)
                                        else:
                                            logging.info("FD BALANCE IS TO LOW PLEASE ADD FUNDS FIRST")
                                    except sqlite3.Error as e:
                                        logging.error("Error occurred while accessing database:", e)
                                    finally:
                                        ccu0.close()
                                else:
                                    exit(0)
                            except ValueError:
                                logging.error("Invalid input. Please enter a valid number.")

                    except sqlite3.Error as e:
                        logging.error("SQLite error:", e)
                    finally:
                        ccu.close()
                else:
                    logging.info("FD is Already Withdrawal....")
            else:
                print("FD account number is not found")

    def check_maturity(self):
            fd_ac = int(input("Enter FD account number"))
            try:
                ccu01 = sqlite3.connect("bank_manage.db")
                cc8 = ccu01.cursor()
                q01 = "SELECT fd_mat_dt FROM fd_accounts WHERE fd_ac_no=?"
                cc8.execute(q01, (fd_ac,))
                r4 = cc8.fetchone()
                logging.info("MATURITY DATE OF FD ACCOUNT %s IS:%s", fd_ac, r4[0])
            except sqlite3.Error as e:
                logging.error(e)

    def extend_date(self):
            fd_ac_no = int(input("Enter the FD Account Number: "))
            if self.check_account(fd_ac_no):
                try:
                    conn = sqlite3.connect("bank_manage.db")
                    cursor = conn.cursor()
                    cursor.execute("SELECT * FROM fd_accounts WHERE fd_ac_no=?", (fd_ac_no,))
                    fd_account = cursor.fetchone()
                    if fd_account:
                        current_mat_date = fd_account[10]
                        days_to_extend = int(input("Enter the number of days to extend the maturity date: "))

                        # Calculate the new maturity date
                        new_mat_date = add_days_to_date(current_mat_date, days_to_extend)
                        # Update the FD account with the new maturity date
                        cursor.execute("UPDATE fd_accounts SET fd_mat_dt=?,fd_days=? WHERE fd_ac_no=?",
                                    (new_mat_date, fd_account[5] + days_to_extend, fd_ac_no))

                        conn.commit()
                        logging.info(
                            "Maturity date extended successfully for FD account {} to {}".format(fd_ac_no, new_mat_date))
                    else:
                        logging.warning("FD account not found with account number {}".format(fd_ac_no))

                except sqlite3.Error as e:
                    logging.error("Error occurred while accessing database:", e)
                finally:
                    if conn:
                        conn.close()
            else:
                print("Account Not found........")

    def renew_fd_ac(self):
            acnn = int(input("Enter Account number to renew FD..."))
            if ac_availability(acnn):
                try:
                    cc0 = sqlite3.connect("bank_manage.db")
                    cc01 = cc0.cursor()
                    cc01.execute("SELECT fd_days,fd_mat_dt,mat_amt,int_rate FROM fd_accounts WHERE fd_ac_no=?", (acnn,))
                    r3 = cc01.fetchone()
                    try:
                        if get_current_date() == r3[1]:
                            logging.info("Maturity date is over Please Renew it....")
                            input("Press Enter to renew account")
                            cc01.execute("UPDATE fd_accounts SET fd_mat_dt=?,renew=renew + ?,balance=?,mat_amt=? WHERE "
                                        "fd_ac_no=?", (
                                            add_days_to_date(r3[1], r3[0]), 1, r3[2],
                                            self.calculate_mat_amt(r3[2], r3[3], r3[0]),
                                            acnn,))
                            cc0.commit()
                            logging.info("FD Renewed successfully")
                            logging.info("MAT-DATE:%S,MAT-AMT:%S", add_days_to_date(r3[1], r3[0]), 1, r3[2],
                                        self.calculate_mat_amt(r3[2], r3[3], r3[0]))
                        else:
                            logging.info("Today is not Maturity date please Renew After Maturity Date")
                    except Exception as e:
                        cc0.rollback()
                        logging.info("Transaction failed")
                except sqlite3.Error as e:
                    logging.info(e)
                finally:
                    if cc0:
                        cc0.close()
def calculate_rd_mat_amt(rd_amt, int_rate, rd_months):
    try:
        # Convert input parameters to appropriate types
        rd_amt = float(rd_amt)
        int_rate = float(int_rate)
        rd_months = int(rd_months)
        
        # Calculate quarterly interest rate and number of quarters
        r = int_rate / 400  # Quarterly interest rate
        n = rd_months // 3  # Number of quarters
        
        # Calculate maturity value using the provided formula
        M = rd_amt * (((1 + r) ** n) - 1) / (1 - (1 + r) ** (-1/3))
        
        # Calculate interest earned
        interest_earned = M - (rd_amt * rd_months)
        
        # Round the results to two decimal places
        M = round(M, 2)
        interest_earned = round(interest_earned, 2)
        
        return M, interest_earned
    except Exception as e:
        print("Error from the calculate_rd_mat_amt:", e)
        return None, None
class RDaccount:
    def create_rd_account(self,user,personal_ac_no,rd_am,rd_month,ir):
        database_name='other_database'
        with transaction.atomic(using=database_name):
            with connections[database_name].cursor() as cursor:
                od = str(get_current_date())
                rd_days = int(rd_month) * 30
                try:
                    customer=PersonalBankAccount.objects.using('other_database').get(account_number=personal_ac_no)
                    res=calculate_rd_mat_amt(rd_am,ir,rd_month)
                    print(res[1],res[0])  
                    rd_mat_amt=res[0]
                    new_account = RDAccountModel.objects.using('other_database').create(
                    customer_name=customer.account_holder_name,
                    opening_date=od,
                    int_rate=ir,
                    rd_months=rd_month,
                    rd_amt=rd_am,
                    int_earned=res[1],
                    personal_ac_no=personal_ac_no,
                    username=user,
                    account_balance=0,  # Assuming initial balance
                    renew=0 ,
                    pre_mature_withdraw=False,
                    mat_amt=rd_mat_amt ,
                    rd_mat_date=add_days_to_date(od, int(rd_days)) ,
                    )
                    new_account.save

                    print("Account Number created sucessfully")
                    print("rd mat amt is",rd_mat_amt)
                    return True  
                except Exception as r:
                    print("the error from the bank_management",r)
                    return False
    def apply_interest(self):
        # Calculate and apply interest to the RD account balance
        pass

    def check_maturity(self):
        # Check if the RD account has matured
        pass

    def withdraw_pre_mature(self):
        # Allow premature withdrawal from the RD account (may not be applicable for RD)
        pass

    def check_interest(self):
        # Check the interest earned on the RD account
        pass

    def check_balance(self):
        # Check the current balance of the RD account
        pass

    def update_account_details(self, monthly_deposit_amount=None, interest_rate=None, duration=None):
        # Update RD account details such as monthly deposit amount, interest rate, or duration
        pass

    def close_account(self):
        # Close the RD account
        pass

    def view_account_statement(self):
        # Provide a statement of account activity
        pass

    def extend_duration(self, additional_duration):
        # Extend the duration of the RD account
        pass


class fd_loan(Fdaccount):
    def __init__(self, an88=None, data=None):
        super().__init__(an88)
        self.data = data

    def calculate_emi(self, amt1, days, annual_interest_rate):
        # Convert days to months
        months = days / 30

        # Convert annual interest rate to monthly interest rate
        monthly_interest_rate = annual_interest_rate / 12 / 100

        # Calculate EMI using the formula
        emi = (amt1 * monthly_interest_rate * (1 + monthly_interest_rate) ** months) / (
                (1 + monthly_interest_rate) ** months - 1)

        return emi

    def create_fd_account(self):
        try:
            res12 = self.get_account_details(self.data)  # Fetch account details
            if res12:
                max_loan_amt = res12[2] * 0.8  # Calculate maximum loan amount (80% of FD amount)
                loan_amt = int(input("Enter the loan amount to give (Enter 80% of the FD_AC): "))
                if loan_amt <= max_loan_amt:  # Check if loan amount is within the allowed limit
                    int_rt = int(input("Enter the interest rate: "))
                    loan_term = int(input("Enter the loan duration in (DAYS): "))
                    try:
                        con11 = sqlite3.connect("bank_manage.db")
                        cur6 = con11.cursor()
                        qu1 = "INSERT INTO FD_Loan (Customer_Name, FD_Amount, Interest_Rate, Loan_Term, Loan_Start_Date, Loan_End_Date, EMI, Outstanding_Principal, Total_Payment_Due, Status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
                        loan_start_date = datetime.now().date()
                        loan_end_date = loan_start_date + timedelta(days=loan_term * 30)  # Assuming 30 days per month
                        emi = self.calculate_emi(loan_amt, loan_term, int_rt)  # Calculate EMI
                        status = "Active"
                        loan_data = (
                            res12[1], res12[2], int_rt, loan_term, loan_start_date, loan_end_date, emi, loan_amt,
                            loan_term * emi, status)
                        cur6.execute(qu1, loan_data)
                        con11.commit()
                        logging.info("Account created successfully.")
                    except Exception as e:
                        con11.rollback()
                        logging.error("Account creation failed:", e)
                else:
                    logging.info(f"Enter an amount less than or equal to {max_loan_amt}.")
            else:
                logging.info("Internal problem occurred.")
        except Exception as ex:
            logging.warning("Database connection failed:", ex)


class rd_loan:
    pass


class personal_loan():
    pass
        


class auto:
    pass


class business_loan:
    pass


class home_laon:
    pass


def authenticate_user(username, entered_password):
    try:
        connection = sqlite3.connect('bank_manage.db')
        cursor = connection.cursor()

        cursor.execute("SELECT hashed_password, bank_name FROM users WHERE username = ?", (username,))
        result = cursor.fetchone()

        connection.close()

        if result:
            stored_hashed_password = result[0]  # Convert hashed password to bytes
            if bcrypt.checkpw(entered_password.encode('utf-8'), stored_hashed_password):
                logging.info("Authentication successful!")
                return True, result[1]
            else:
                logging.warning("Authentication failed.")
                return False
        else:
            logging.warning("User not found.")
            return False
    except TypeError as e:
        logging.critical("An error occurred during authentication: %s", e)
        return None

